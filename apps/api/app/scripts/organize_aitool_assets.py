from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook
from app.services.logo_assets import LOGO_SOURCE_IMPORTED, resolve_logo_status


DEFAULT_TARGET_DIR = (
    Path(__file__).resolve().parents[4]
    / "archive"
    / "drawer"
    / "tooling-assets"
    / "apigetxlsx"
    / "aitool"
)

PLACEHOLDER_LOGO_REFS = {"image.png"}
EMPTY_LIKE_VALUES = {"", "-", "n/a", "na", "none", "null", "未显示", "未提供", "无"}
TOKEN_STOP_WORDS = {
    "ai",
    "app",
    "co",
    "com",
    "company",
    "cn",
    "inc",
    "io",
    "logo",
    "net",
    "org",
    "studio",
    "tech",
    "the",
    "tools",
    "web",
    "www",
}


@dataclass
class SheetSummary:
    title: str
    rows: int
    columns: int
    headers: list[str]


@dataclass
class LogoEntry:
    file_name: str
    stem: str
    normalized: str
    tokens: set[str]


def ensure_layout(root: Path) -> dict[str, Path]:
    source_dir = root / "source"
    manifests_dir = root / "manifests"
    spreadsheets_dir = source_dir / "spreadsheets"
    logos_dir = source_dir / "logos"
    archives_dir = source_dir / "archives"

    for directory in (source_dir, manifests_dir, spreadsheets_dir, logos_dir, archives_dir):
        directory.mkdir(parents=True, exist_ok=True)

    return {
        "root": root,
        "source": source_dir,
        "manifests": manifests_dir,
        "spreadsheets": spreadsheets_dir,
        "logos": logos_dir,
        "archives": archives_dir,
    }


def move_if_needed(source: Path, destination_dir: Path) -> Path:
    destination = destination_dir / source.name
    if source.resolve() == destination.resolve():
        return destination
    if destination.exists():
        return destination
    source.rename(destination)
    return destination


def reorganize_sources(root: Path, layout: dict[str, Path]) -> None:
    for item in list(root.iterdir()):
        if item.name in {"source", "manifests", "README.md"}:
            continue
        if item.is_file() and item.suffix.lower() == ".xlsx":
            move_if_needed(item, layout["spreadsheets"])
        elif item.is_file() and item.suffix.lower() in {".zip", ".base"}:
            move_if_needed(item, layout["archives"])
        elif item.is_dir():
            move_if_needed(item, layout["logos"])


def find_workbook(spreadsheets_dir: Path) -> Path:
    candidates = sorted(spreadsheets_dir.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No .xlsx workbook found in {spreadsheets_dir}")
    return candidates[0]


def find_logo_dir(logos_dir: Path) -> Path:
    directories = [path for path in logos_dir.iterdir() if path.is_dir()]
    if directories:
        return directories[0]
    return logos_dir


def summarize_workbook(workbook_path: Path) -> tuple[list[SheetSummary], list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_summaries: list[SheetSummary] = []
    tool_rows: list[dict[str, Any]] = []

    for index, worksheet in enumerate(workbook.worksheets):
        headers_row = next(worksheet.iter_rows(values_only=True), ())
        headers = [normalize_text(value) for value in headers_row]
        sheet_summaries.append(
            SheetSummary(
                title=worksheet.title,
                rows=worksheet.max_row,
                columns=worksheet.max_column,
                headers=headers,
            )
        )

        if index != 0:
            continue

        for row in worksheet.iter_rows(values_only=True, min_row=2):
            if not any(row):
                continue
            tool_rows.append(
                {
                    "name": normalize_text(row[0]),
                    "logo_ref": normalize_text(row[1]),
                    "developer": normalize_text(row[2]),
                    "country": normalize_text(row[3]),
                    "city": normalize_text(row[4]),
                    "subtitle": normalize_text(row[5]),
                    "homepage_screenshot": normalize_text(row[6]),
                    "price": normalize_text(row[7]),
                    "price_screenshot": normalize_text(row[8]),
                    "special_tags": normalize_text(row[9]),
                    "tags": normalize_text(row[10]),
                    "platforms": normalize_text(row[11]),
                    "vpn_required": normalize_text(row[12]),
                    "url": normalize_text(row[13]),
                    "remark": normalize_text(row[14]),
                    "detail_page": normalize_text(row[15]),
                    "parent_record": normalize_text(row[16]),
                }
            )

    return sheet_summaries, tool_rows


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_logo_catalog(logo_dir: Path) -> tuple[list[Path], dict[str, Path], list[LogoEntry]]:
    logo_files = sorted(path for path in logo_dir.iterdir() if path.is_file())
    logo_lookup: dict[str, Path] = {}
    logo_entries: list[LogoEntry] = []

    for path in logo_files:
        logo_lookup[path.name] = path
        logo_lookup.setdefault(path.name.lower(), path)
        logo_entries.append(
            LogoEntry(
                file_name=path.name,
                stem=path.stem,
                normalized=normalize_identifier(path.stem),
                tokens=tokenize_identity(path.stem),
            )
        )

    return logo_files, logo_lookup, logo_entries


def build_tool_logo_report(
    tool_rows: list[dict[str, Any]],
    logo_dir: Path,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    logo_files, logo_lookup, logo_entries = build_logo_catalog(logo_dir)
    exact_logo_counter = Counter(
        matched_logo
        for row in tool_rows
        for _, matched_logo, _ in [resolve_logo_match(row, logo_lookup, logo_entries)]
        if matched_logo
    )

    row_reports: list[dict[str, Any]] = []
    status_counter: Counter[str] = Counter()
    risk_counter: Counter[str] = Counter()

    for row in tool_rows:
        status, matched_logo, matched_score = resolve_logo_match(row, logo_lookup, logo_entries)

        risk = assess_logo_risk(
            row=row,
            status=status,
            matched_logo=matched_logo,
            logo_entries=logo_entries,
            exact_logo_counter=exact_logo_counter,
            matched_score=matched_score,
        )

        report_row = {
            **row,
            "logo_status": status,
            "matched_logo": matched_logo,
            "logo_risk_score": risk["score"],
            "logo_risk_level": risk["level"],
            "logo_risk_reasons": "|".join(risk["reasons"]),
            "expected_identity_tokens": "|".join(sorted(risk["expected_tokens"])),
            "matched_logo_tokens": "|".join(sorted(risk["matched_tokens"])),
            "better_logo_candidate": risk["better_logo_candidate"],
            "better_logo_candidate_score": risk["better_logo_candidate_score"],
        }
        row_reports.append(report_row)
        status_counter[status] += 1
        risk_counter[risk["level"]] += 1

    summary = {
        "tool_rows": len(tool_rows),
        "logo_files": len(logo_files),
        "placeholder_logo_refs": status_counter["placeholder"],
        "missing_logo_refs": status_counter["missing"],
        "exact_logo_matches": status_counter["exact_match"],
        "auto_logo_matches": status_counter["auto_match"],
        "unresolved_non_placeholder_logo_refs": status_counter["unresolved"],
        "high_risk_logo_rows": risk_counter["high"] + risk_counter["critical"],
        "risk_level_breakdown": dict(sorted(risk_counter.items())),
        "logo_extensions": count_extensions(logo_files),
    }
    return summary, row_reports


def resolve_logo_match(
    row: dict[str, Any],
    logo_lookup: dict[str, Path],
    logo_entries: list[LogoEntry],
) -> tuple[str, str, float]:
    logo_ref = normalize_text(row.get("logo_ref", ""))
    matched_path = logo_lookup.get(logo_ref) or logo_lookup.get(logo_ref.lower())
    if matched_path:
        return "exact_match", matched_path.name, 1.0

    best_candidate, best_score = find_better_logo_candidate(
        identity_candidates=build_identity_candidates(row),
        current_logo_name="",
        current_similarity=0.0,
        logo_entries=logo_entries,
    )
    if best_candidate:
        return "auto_match", best_candidate, best_score

    if not logo_ref:
        return "missing", "", 0.0
    if logo_ref.lower() in PLACEHOLDER_LOGO_REFS:
        return "placeholder", "", 0.0
    return "unresolved", "", 0.0


def assess_logo_risk(
    *,
    row: dict[str, Any],
    status: str,
    matched_logo: str,
    logo_entries: list[LogoEntry],
    exact_logo_counter: Counter[str],
    matched_score: float,
) -> dict[str, Any]:
    expected_tokens = build_expected_identity_tokens(row)
    matched_entry = next((item for item in logo_entries if item.file_name == matched_logo), None)
    matched_tokens = matched_entry.tokens if matched_entry else set()
    reasons: list[str] = []
    score = 0

    if status == "missing":
        return {
            "score": 100,
            "level": "critical",
            "reasons": ["missing_logo_ref"],
            "expected_tokens": expected_tokens,
            "matched_tokens": matched_tokens,
            "better_logo_candidate": "",
            "better_logo_candidate_score": 0.0,
        }

    if status == "placeholder":
        return {
            "score": 95,
            "level": "critical",
            "reasons": ["placeholder_logo_ref"],
            "expected_tokens": expected_tokens,
            "matched_tokens": matched_tokens,
            "better_logo_candidate": "",
            "better_logo_candidate_score": 0.0,
        }

    if status == "unresolved":
        return {
            "score": 90,
            "level": "critical",
            "reasons": ["logo_file_not_found"],
            "expected_tokens": expected_tokens,
            "matched_tokens": matched_tokens,
            "better_logo_candidate": "",
            "better_logo_candidate_score": 0.0,
        }

    identity_candidates = build_identity_candidates(row)
    current_similarity = best_similarity(identity_candidates, matched_entry.normalized if matched_entry else "")
    overlap = expected_tokens & matched_tokens

    if matched_logo and exact_logo_counter[matched_logo] > 1:
        score += 25
        reasons.append("logo_file_reused_by_multiple_rows")

    if status == "auto_match":
        reasons.append("auto_matched_from_logo_pack")

    if expected_tokens and not overlap:
        score += 35
        reasons.append("no_identity_token_overlap")

    if current_similarity < 0.28:
        score += 25
        reasons.append("low_name_similarity")

    domain_token = primary_domain_token(row.get("url", ""))
    if domain_token and domain_token not in matched_tokens and current_similarity < 0.45:
        score += 15
        reasons.append("domain_token_missing_from_logo_name")

    if status == "auto_match" and matched_score >= 0.95:
        score = max(score - 15, 0)
    elif status == "auto_match" and matched_score >= 0.85:
        score = max(score - 5, 0)

    better_logo_candidate, better_logo_candidate_score = find_better_logo_candidate(
        identity_candidates=identity_candidates,
        current_logo_name=matched_logo,
        current_similarity=current_similarity,
        logo_entries=logo_entries,
    )
    if better_logo_candidate:
        score += 40
        reasons.append("better_logo_candidate_exists")

    score = min(score, 100)
    level = risk_level_from_score(score)
    return {
        "score": score,
        "level": level,
        "reasons": reasons,
        "expected_tokens": expected_tokens,
        "matched_tokens": matched_tokens,
        "better_logo_candidate": better_logo_candidate,
        "better_logo_candidate_score": round(better_logo_candidate_score, 3),
    }


def build_identity_candidates(row: dict[str, Any]) -> list[str]:
    values = [
        normalize_identifier(row.get("name", "")),
        normalize_identifier(row.get("developer", "")),
        normalize_identifier(primary_domain_token(row.get("url", ""))),
    ]
    return [value for value in values if value]


def build_expected_identity_tokens(row: dict[str, Any]) -> set[str]:
    tokens = set()
    tokens.update(tokenize_identity(row.get("name", "")))
    tokens.update(tokenize_identity(row.get("developer", "")))
    domain_token = primary_domain_token(row.get("url", ""))
    if domain_token:
        tokens.add(domain_token)
    return tokens


def find_better_logo_candidate(
    *,
    identity_candidates: list[str],
    current_logo_name: str,
    current_similarity: float,
    logo_entries: list[LogoEntry],
) -> tuple[str, float]:
    if not identity_candidates:
        return "", 0.0

    best_candidate = ""
    best_score = 0.0

    for entry in logo_entries:
        if entry.file_name == current_logo_name:
            continue
        score = best_similarity(identity_candidates, entry.normalized)
        if score > best_score:
            best_score = score
            best_candidate = entry.file_name

    if best_score >= 0.78 and best_score - current_similarity >= 0.25:
        return best_candidate, best_score
    return "", 0.0


def best_similarity(candidates: list[str], target: str) -> float:
    if not candidates or not target:
        return 0.0
    return max(SequenceMatcher(None, candidate, target).ratio() for candidate in candidates)


def risk_level_from_score(score: int) -> str:
    if score >= 90:
        return "critical"
    if score >= 60:
        return "high"
    if score >= 30:
        return "medium"
    return "low"


def normalize_identifier(value: str) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"https?://", "", text)
    text = re.sub(r"www\.", "", text)
    text = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text)
    return text


def tokenize_identity(value: str) -> set[str]:
    text = normalize_text(value).lower()
    parts = re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]+", text)
    tokens: set[str] = set()
    for part in parts:
        if part in TOKEN_STOP_WORDS:
            continue
        if re.fullmatch(r"[a-z]", part):
            continue
        tokens.add(part)
    return tokens


def primary_domain_token(raw_url: str) -> str:
    sanitized = sanitize_url(raw_url)
    parsed = urlparse(sanitized)
    host = parsed.netloc.lower().replace("www.", "")
    if not host:
        return ""
    pieces = [piece for piece in host.split(".") if piece and piece not in {"com", "cn", "net", "org", "io", "ai"}]
    if not pieces:
        return ""
    return normalize_identifier(pieces[0])


def build_import_payload(tool_rows: list[dict[str, Any]], logo_dir: Path) -> dict[str, Any]:
    categories: dict[str, dict[str, str]] = {}
    tags: dict[str, dict[str, str]] = {}
    tools: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []
    tool_tags: list[dict[str, str]] = []

    _, logo_lookup, logo_entries = build_logo_catalog(logo_dir)
    seen_slugs: dict[str, int] = {}
    today = date.today().isoformat()

    for row in tool_rows:
        if not row["name"]:
            continue

        slug = make_slug(row["name"], row["url"], seen_slugs)
        category_name = derive_category(row)
        tag_names = split_tags(row["tags"])
        special_tag_names = split_tags(row["special_tags"])
        all_tags = dedupe_preserve_order(tag_names + special_tag_names)

        if category_name:
            categories.setdefault(
                category_name,
                {
                    "slug": slugify_text(category_name),
                    "name": category_name,
                    "description": f"Imported from workbook category: {category_name}",
                },
            )

        for tag_name in all_tags:
            tags.setdefault(tag_name, {"name": tag_name})
            tool_tags.append({"tool_slug": slug, "tag_name": tag_name})

        logo_path = derive_logo_path(row, logo_lookup, logo_entries, logo_dir)
        description = row["subtitle"] or row["remark"] or row["name"]
        editor_comment_parts = [
            row["developer"] and f"Developer: {row['developer']}",
            row["country"] and f"Country: {row['country']}",
            row["city"] and f"City: {row['city']}",
            row["price"] and f"Price: {row['price']}",
            row["platforms"] and f"Platforms: {row['platforms']}",
            row["vpn_required"] and f"VPN required: {row['vpn_required']}",
            row["remark"] and f"Remark: {row['remark']}",
        ]
        editor_comment = " | ".join(part for part in editor_comment_parts if part)

        tools.append(
            {
                "slug": slug,
                "name": row["name"],
                "category_name": category_name,
                "summary": truncate_text(description, 120),
                "description": description,
                "editor_comment": editor_comment or f"Imported from workbook row for {row['name']}",
                "official_url": sanitize_url(row["url"]),
                "logo_path": logo_path,
                "logo_status": resolve_logo_status(logo_path),
                "logo_source": LOGO_SOURCE_IMPORTED,
                "score": 0.0,
                "status": "draft",
                "featured": False,
                "created_on": today,
                "last_verified_at": today,
                "import_meta": {
                    "developer": row["developer"],
                    "country": row["country"],
                    "city": row["city"],
                    "price": row["price"],
                    "price_screenshot": row["price_screenshot"],
                    "homepage_screenshot": row["homepage_screenshot"],
                    "platforms": row["platforms"],
                    "vpn_required": row["vpn_required"],
                    "detail_page": row["detail_page"],
                    "parent_record": row["parent_record"],
                    "logo_ref": row["logo_ref"],
                },
            }
        )

        if row["url"]:
            sources.append(
                {
                    "tool_slug": slug,
                    "source_type": "workbook_link",
                    "source_url": sanitize_url(row["url"]),
                }
            )

    return {
        "tools": tools,
        "categories": sorted(categories.values(), key=lambda item: item["name"]),
        "tags": sorted(tags.values(), key=lambda item: item["name"]),
        "tool_tags": tool_tags,
        "sources": sources,
    }


def derive_category(row: dict[str, Any]) -> str:
    primary_tags = split_tags(row["tags"])
    if primary_tags:
        return primary_tags[0]
    special_tags = split_tags(row["special_tags"])
    if special_tags:
        return special_tags[0]
    return "uncategorized"


def split_tags(raw: str) -> list[str]:
    value = normalize_text(raw)
    if is_empty_like(value):
        return []
    items = re.split(r"[,，、/|;；]+", value)
    return [item.strip() for item in items if item.strip() and not is_empty_like(item)]


def is_empty_like(value: str) -> bool:
    return normalize_text(value).lower() in EMPTY_LIKE_VALUES


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def slugify_text(value: str) -> str:
    lowered = value.strip().lower()
    lowered = re.sub(r"https?://", "", lowered)
    lowered = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", lowered)
    lowered = lowered.strip("-")
    return lowered or "item"


def slug_from_url(raw_url: str) -> str:
    parsed = urlparse(sanitize_url(raw_url))
    host = parsed.netloc.lower().replace("www.", "")
    path = parsed.path.strip("/").lower()
    path = re.sub(r"[^a-z0-9/-]+", "-", path)
    host = re.sub(r"[^a-z0-9.-]+", "-", host)
    pieces = [piece for piece in [host.split(".")[0] if host else "", path.split("/")[0] if path else ""] if piece]
    candidate = "-".join(pieces).strip("-")
    return candidate or ""


def make_slug(name: str, raw_url: str, seen_slugs: dict[str, int]) -> str:
    base = slug_from_url(raw_url) or slugify_text(name)
    count = seen_slugs.get(base, 0)
    seen_slugs[base] = count + 1
    if count == 0:
        return base
    return f"{base}-{count + 1}"


def sanitize_url(raw_url: str) -> str:
    if not raw_url:
        return ""
    if raw_url.startswith(("http://", "https://")):
        return raw_url
    if "." in raw_url and " " not in raw_url:
        return f"https://{raw_url}"
    return raw_url


def derive_logo_path(
    row: dict[str, Any],
    logo_lookup: dict[str, Path],
    logo_entries: list[LogoEntry],
    logo_dir: Path,
) -> str | None:
    _, matched_logo, _ = resolve_logo_match(row, logo_lookup, logo_entries)
    if not matched_logo:
        return None
    matched = logo_lookup.get(matched_logo) or logo_lookup.get(matched_logo.lower())
    if not matched:
        return None
    assets_root = logo_dir.parents[3]
    return str(matched.relative_to(assets_root)).replace("\\", "/")


def truncate_text(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def count_extensions(paths: list[Path]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for path in paths:
        suffix = path.suffix.lower() or "<none>"
        counts[suffix] = counts.get(suffix, 0) + 1
    return dict(sorted(counts.items()))


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Organize AI tool workbook and logo assets into a stable ingest layout."
    )
    parser.add_argument(
        "--target-dir",
        type=Path,
        default=DEFAULT_TARGET_DIR,
        help="Target aitool asset directory.",
    )
    args = parser.parse_args()

    root = args.target_dir.resolve()
    if not root.exists():
        raise SystemExit(f"Target directory does not exist: {root}")

    layout = ensure_layout(root)
    reorganize_sources(root, layout)

    workbook_path = find_workbook(layout["spreadsheets"])
    logo_dir = find_logo_dir(layout["logos"])

    sheet_summaries, tool_rows = summarize_workbook(workbook_path)
    logo_summary, tool_logo_rows = build_tool_logo_report(tool_rows, logo_dir)
    import_payload = build_import_payload(tool_rows, logo_dir)

    summary = {
        "target_dir": str(root),
        "workbook_path": str(workbook_path),
        "logo_dir": str(logo_dir),
        "sheets": [asdict(item) for item in sheet_summaries],
        "tool_logo_summary": logo_summary,
        "db_import_summary": {
            "tools": len(import_payload["tools"]),
            "categories": len(import_payload["categories"]),
            "tags": len(import_payload["tags"]),
            "tool_tags": len(import_payload["tool_tags"]),
            "sources": len(import_payload["sources"]),
        },
    }

    manifests_dir = layout["manifests"]
    write_json(manifests_dir / "asset_summary.json", summary)
    write_json(manifests_dir / "db_import_payload.json", import_payload)
    write_csv(manifests_dir / "tool_logo_report.csv", tool_logo_rows)
    write_csv(
        manifests_dir / "unresolved_logo_refs.csv",
        [row for row in tool_logo_rows if row["logo_status"] == "unresolved"],
    )
    write_csv(
        manifests_dir / "placeholder_logo_refs.csv",
        [row for row in tool_logo_rows if row["logo_status"] == "placeholder"],
    )
    write_csv(
        manifests_dir / "suspicious_logo_matches.csv",
        [row for row in tool_logo_rows if row["logo_risk_level"] in {"high", "critical"}],
    )

    print(f"Organized assets under: {root}")
    print(f"Workbook: {workbook_path.name}")
    print(f"Logo directory: {logo_dir.name}")
    print(f"Tool rows: {logo_summary['tool_rows']}")
    print(f"Logo files: {logo_summary['logo_files']}")
    print(f"Placeholder logo refs: {logo_summary['placeholder_logo_refs']}")
    print(f"Unresolved non-placeholder logo refs: {logo_summary['unresolved_non_placeholder_logo_refs']}")
    print(f"High-risk logo rows: {logo_summary['high_risk_logo_rows']}")
    print(f"DB import tools: {len(import_payload['tools'])}")


if __name__ == "__main__":
    main()
